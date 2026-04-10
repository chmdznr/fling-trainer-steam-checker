"""
FLiNG Trainer + Steam Deck Compatibility Checker

Scrapes FLiNG Trainer website for games with trainers (year >= MIN_YEAR),
then checks Steam Store for Deck compatibility, price, and ratings.
Outputs results to a formatted Excel file.

Uses a local JSON cache so subsequent runs only fetch new trainers.
Price/discount data is always refreshed since it changes frequently.

Usage:
    python fling_steam_checker.py                    # Default (ID/IDR, year 2024+)
    python fling_steam_checker.py --country US        # USD pricing
    python fling_steam_checker.py --country JP        # JPY pricing
    python fling_steam_checker.py --year 2023         # Include older trainers
    python fling_steam_checker.py --no-cache          # Fresh run, ignore cache
    python fling_steam_checker.py --help              # Show all options
"""

import argparse
import sys
import requests
from bs4 import BeautifulSoup
import re
import time
import json
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# Force unbuffered output so progress is visible in real-time
_builtin_print = print
def print(*args, **kwargs):
    kwargs.setdefault("flush", True)
    _builtin_print(*args, **kwargs)

# ─── Currency Map ─────────────────────────────────────────────

CURRENCY_MAP = {
    "AR": {"symbol": "ARS$", "code": "ARS", "decimal": True,  "sep": ","},
    "AU": {"symbol": "A$",   "code": "AUD", "decimal": True,  "sep": ","},
    "BR": {"symbol": "R$",   "code": "BRL", "decimal": True,  "sep": "."},
    "CA": {"symbol": "C$",   "code": "CAD", "decimal": True,  "sep": ","},
    "CN": {"symbol": "¥",    "code": "CNY", "decimal": True,  "sep": ","},
    "EU": {"symbol": "€",    "code": "EUR", "decimal": True,  "sep": ","},
    "GB": {"symbol": "£",    "code": "GBP", "decimal": True,  "sep": ","},
    "ID": {"symbol": "Rp",   "code": "IDR", "decimal": False, "sep": "."},
    "IN": {"symbol": "₹",    "code": "INR", "decimal": True,  "sep": ","},
    "JP": {"symbol": "¥",    "code": "JPY", "decimal": False, "sep": ","},
    "KR": {"symbol": "₩",    "code": "KRW", "decimal": False, "sep": ","},
    "MX": {"symbol": "Mex$", "code": "MXN", "decimal": True,  "sep": ","},
    "NO": {"symbol": "kr",   "code": "NOK", "decimal": True,  "sep": "."},
    "NZ": {"symbol": "NZ$",  "code": "NZD", "decimal": True,  "sep": ","},
    "PH": {"symbol": "₱",    "code": "PHP", "decimal": True,  "sep": ","},
    "PL": {"symbol": "zł",   "code": "PLN", "decimal": True,  "sep": " "},
    "RU": {"symbol": "₽",    "code": "RUB", "decimal": True,  "sep": " "},
    "SA": {"symbol": "SR",   "code": "SAR", "decimal": True,  "sep": ","},
    "SE": {"symbol": "kr",   "code": "SEK", "decimal": True,  "sep": "."},
    "SG": {"symbol": "S$",   "code": "SGD", "decimal": True,  "sep": ","},
    "TH": {"symbol": "฿",    "code": "THB", "decimal": True,  "sep": ","},
    "TR": {"symbol": "₺",    "code": "TRY", "decimal": True,  "sep": "."},
    "TW": {"symbol": "NT$",  "code": "TWD", "decimal": False, "sep": ","},
    "UA": {"symbol": "₴",    "code": "UAH", "decimal": True,  "sep": " "},
    "US": {"symbol": "$",    "code": "USD", "decimal": True,  "sep": ","},
    "ZA": {"symbol": "R",    "code": "ZAR", "decimal": True,  "sep": " "},
}

DEFAULT_CURRENCY = {"symbol": "$", "code": "USD", "decimal": True, "sep": ","}

# ─── Steam URLs ────────────────────────────────────────────────

FLING_BASE_URL = "https://flingtrainer.com/category/trainer/page/{page}/"
FLING_FIRST_PAGE = "https://flingtrainer.com/category/trainer/"
STEAM_SEARCH_URL = "https://store.steampowered.com/api/storesearch/"
STEAM_APPDETAILS_URL = "https://store.steampowered.com/api/appdetails"
STEAM_REVIEWS_URL = "https://store.steampowered.com/appreviews/{appid}"
STEAM_DECK_URL = "https://store.steampowered.com/saleaction/ajaxgetdeckappcompatibilityreport"

DECK_COMPAT_MAP = {
    0: "Unknown",
    1: "Unsupported",
    2: "Playable",
    3: "Verified",
}

# ─── Config ────────────────────────────────────────────────────

@dataclass
class Config:
    min_year: int = 2024
    country_code: str = "ID"
    request_delay: float = 1.5
    max_workers: int = 4
    max_retries: int = 3
    cache_price_ttl_hours: int = 24
    cache_full_ttl_days: int = 7
    output_dir: Path = None  # default: current directory
    no_cache: bool = False
    verbose: bool = False
    # derived:
    currency: dict = field(default_factory=dict)

    def __post_init__(self):
        if self.output_dir is None:
            self.output_dir = Path.cwd()
        self.currency = CURRENCY_MAP.get(self.country_code, DEFAULT_CURRENCY)
        self.cache_path = self.output_dir / "fling_steam_cache.json"

    @property
    def currency_code(self) -> str:
        return self.currency["code"]

    @property
    def currency_symbol(self) -> str:
        return self.currency["symbol"]

    @property
    def price_number_format(self) -> str:
        """Excel number format string based on currency."""
        if self.currency["decimal"]:
            return '#,##0.00'
        return '#,##0'

    @property
    def price_header(self) -> str:
        return f"Price ({self.currency_code})"


def build_session() -> requests.Session:
    """Create a requests session with browser-like headers."""
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return session


# ─── Cache ────────────────────────────────────────────────────

def load_cache(config: Config) -> dict:
    """Load cache from JSON file. Keyed by trainer_url."""
    if config.no_cache or not config.cache_path.exists():
        return {}
    try:
        with open(config.cache_path) as f:
            data = json.load(f)
        print(f"  Loaded cache: {len(data)} entries from {config.cache_path.name}")
        return data
    except (json.JSONDecodeError, OSError):
        return {}


def save_cache(cache: dict, config: Config):
    """Save cache to JSON file."""
    with open(config.cache_path, "w") as f:
        json.dump(cache, f, indent=2, default=str)
    print(f"  Cache saved: {len(cache)} entries to {config.cache_path.name}")


# ─── FLiNG Scraping ─────────────────────────────────────────

def scrape_fling_trainers(cache: dict, config: Config) -> tuple[list[dict], list[dict]]:
    """
    Scrape FLiNG Trainer category pages for trainers from min_year onwards.
    Stops when it hits a trainer_url already in cache.

    Returns:
        new_trainers: trainers not in cache (need Steam lookup)
        cached_trainers: trainers found in cache (skip Steam lookup)
    """
    new_trainers = []
    cached_trainers = []
    cached_urls = set(cache.keys())
    page = 1
    hit_cache = False

    pbar = tqdm(desc="  Scraping FLiNG", unit="page")

    while True:
        url = FLING_FIRST_PAGE if page == 1 else FLING_BASE_URL.format(page=page)

        try:
            resp = SESSION.get(url, timeout=15)
            if resp.status_code == 404:
                tqdm.write(f"  Page {page} not found, stopping.")
                break
            resp.raise_for_status()
        except requests.RequestException as e:
            tqdm.write(f"  Error fetching page {page}: {e}")
            break

        soup = BeautifulSoup(resp.text, "html.parser")
        articles = soup.select("article")

        if not articles:
            tqdm.write(f"  No articles found on page {page}, stopping.")
            break

        stop_scraping = False
        for article in articles:
            year_el = article.select_one(".post-details-year")
            month_el = article.select_one(".post-details-month")
            day_el = article.select_one(".post-details-day")

            if not all([year_el, month_el, day_el]):
                continue

            year = int(year_el.text.strip())
            if year < config.min_year:
                stop_scraping = True
                break

            title_el = article.select_one(".post-title a")
            if not title_el:
                continue

            trainer_url = title_el.get("href", "")

            if trainer_url in cached_urls:
                hit_cache = True
                cached_trainers.append(cache[trainer_url])
                continue

            date_str = f"{day_el.text.strip()} {month_el.text.strip()} {year}"
            try:
                date = datetime.strptime(date_str, "%d %b %Y")
            except ValueError:
                date = None

            title = title_el.text.strip()

            entry_el = article.select_one(".entry p")
            entry_text = entry_el.get_text(strip=True) if entry_el else ""

            options_match = re.search(r"(\d+)\s*Options?", entry_text)
            options_count = int(options_match.group(1)) if options_match else None

            version_match = re.search(r"Game Version:\s*(.+?)(?:\s*·|$)", entry_text)
            game_version = version_match.group(1).strip() if version_match else ""

            updated_match = re.search(r"Last Updated:\s*([\d.]+)", entry_text)
            last_updated = updated_match.group(1).strip() if updated_match else ""

            game_name = re.sub(r"\s+Trainer\s*$", "", title).strip()

            new_trainers.append({
                "game_name": game_name,
                "trainer_title": title,
                "trainer_url": trainer_url,
                "trainer_date": date.isoformat() if date else None,
                "trainer_date_str": date_str,
                "options_count": options_count,
                "game_version": game_version,
                "last_updated": last_updated,
            })

        pbar.update(1)

        if stop_scraping or hit_cache:
            if hit_cache:
                for url, entry in cache.items():
                    if url not in {t["trainer_url"] for t in cached_trainers} and \
                       url not in {t["trainer_url"] for t in new_trainers}:
                        cached_trainers.append(entry)
                tqdm.write(f"  Hit cached data on page {page}, stopping scrape.")
            if stop_scraping:
                tqdm.write(f"  Reached year < {config.min_year} on page {page}, stopping.")
            break

        page += 1
        time.sleep(0.5)

    pbar.close()

    print(f"\n  New trainers: {len(new_trainers)}")
    print(f"  From cache:   {len(cached_trainers)}")
    print(f"  Total:        {len(new_trainers) + len(cached_trainers)}")
    return new_trainers, cached_trainers


# ─── Steam API (with retry) ────────────────────────────────────

def steam_request(url: str, params: dict, config: Config, timeout: int = 10) -> requests.Response | None:
    """Make a Steam API request with exponential backoff retry.

    Retries on transient errors (connection errors, timeouts, 5xx).
    Returns None on permanent failure (4xx) or exhausted retries.
    """
    for attempt in range(config.max_retries):
        try:
            resp = SESSION.get(url, params=params, timeout=timeout)
            resp.raise_for_status()
            return resp
        except requests.exceptions.ConnectionError:
            pass
        except requests.exceptions.Timeout:
            pass
        except requests.exceptions.HTTPError:
            if resp.status_code >= 500:
                pass  # retry on server errors
            else:
                return None  # client error, don't retry
        except requests.RequestException:
            return None

        if attempt < config.max_retries - 1:
            wait = 2 ** (attempt + 1)  # 2s, 4s, 8s
            time.sleep(wait)

    return None


def search_steam_appid(game_name: str, config: Config) -> dict | None:
    """Search Steam Store for a game and return best match."""
    params = {"term": game_name, "l": "english", "cc": config.country_code}
    resp = steam_request(STEAM_SEARCH_URL, params=params, config=config)
    if resp is None:
        return None

    try:
        data = resp.json()
    except json.JSONDecodeError:
        return None

    items = data.get("items", [])
    if not items:
        return None

    name_lower = game_name.lower()
    for item in items:
        if item.get("name", "").lower() == name_lower:
            return {"appid": item["id"], "name": item["name"]}
    return {"appid": items[0]["id"], "name": items[0]["name"]}


def get_steam_app_details(appid: int, config: Config) -> dict | None:
    """Get app details from Steam Store API."""
    params = {"appids": appid, "cc": config.country_code, "l": "english"}
    resp = steam_request(STEAM_APPDETAILS_URL, params=params, config=config)
    if resp is None:
        return None

    try:
        data = resp.json()
    except json.JSONDecodeError:
        return None

    app_data = data.get(str(appid), {})
    if not app_data.get("success"):
        return None
    return app_data.get("data")


def get_steam_deck_compat(appid: int, config: Config) -> str:
    """Check Steam Deck compatibility for an app."""
    params = {"nAppID": appid}
    resp = steam_request(STEAM_DECK_URL, params=params, config=config)
    if resp is None:
        return "Unknown"

    try:
        data = resp.json()
    except json.JSONDecodeError:
        return "Unknown"

    resolved_category = data.get("results", {}).get("resolved_category", 0)
    return DECK_COMPAT_MAP.get(resolved_category, "Unknown")


def get_steam_reviews(appid: int, config: Config) -> dict:
    """Get review summary for an app."""
    params = {
        "json": 1, "language": "all", "purchase_type": "all",
        "num_per_page": 0, "review_type": "all",
    }
    url = STEAM_REVIEWS_URL.format(appid=appid)
    resp = steam_request(url, params=params, config=config)
    if resp is None:
        return {}

    try:
        data = resp.json()
    except json.JSONDecodeError:
        return {}

    summary = data.get("query_summary", {})
    total = summary.get("total_reviews", 0)
    positive = summary.get("total_positive", 0)
    pct = round(positive / total * 100, 1) if total > 0 else 0

    return {
        "total_reviews": total,
        "positive_pct": pct,
        "review_desc": summary.get("review_score_desc", "No Reviews"),
    }


def extract_price_info(app_data: dict, config: Config) -> dict:
    """Extract price and discount info from app details."""
    if app_data.get("is_free"):
        return {
            "price": "Free", "price_idr": 0, "original_price_idr": 0,
            "discount_pct": 0, "on_sale": False,
        }

    price_overview = app_data.get("price_overview")
    if not price_overview:
        return {
            "price": "N/A", "price_idr": None, "original_price_idr": None,
            "discount_pct": 0, "on_sale": False,
        }

    return {
        "price": price_overview.get("final_formatted", "N/A"),
        "price_idr": price_overview.get("final", 0) / 100,
        "original_price_idr": price_overview.get("initial", 0) / 100,
        "discount_pct": price_overview.get("discount_percent", 0),
        "on_sale": price_overview.get("discount_percent", 0) > 0,
    }


# ─── Processing ──────────────────────────────────────────────

def _process_single_new_trainer(trainer: dict, config: Config) -> dict:
    """Process a single new trainer: full Steam lookup (search + deck + details + reviews)."""
    game_name = trainer["game_name"]
    now = datetime.now().isoformat()

    steam_match = search_steam_appid(game_name, config)
    time.sleep(config.request_delay)

    if not steam_match:
        tqdm.write(f"  ⚠ {game_name} — not found on Steam")
        return {
            **trainer,
            "steam_appid": None, "steam_name": None, "steam_url": None,
            "deck_compat": "Not Found",
            "price": "N/A", "price_idr": None, "original_price_idr": None,
            "discount_pct": 0, "on_sale": False,
            "total_reviews": 0, "positive_pct": 0, "review_desc": "Not Found",
            "last_fetched": now,
        }

    appid = steam_match["appid"]
    steam_name = steam_match["name"]
    tqdm.write(f"  ✓ {game_name} → {steam_name} (ID: {appid})")

    deck_compat = get_steam_deck_compat(appid, config)
    time.sleep(config.request_delay)

    app_details = get_steam_app_details(appid, config)
    time.sleep(config.request_delay)
    price_info = extract_price_info(app_details, config) if app_details else {
        "price": "N/A", "price_idr": None, "original_price_idr": None,
        "discount_pct": 0, "on_sale": False,
    }

    reviews = get_steam_reviews(appid, config)
    time.sleep(config.request_delay)

    if price_info.get("on_sale"):
        tqdm.write(f"  💰 ON SALE: {game_name} — {price_info['price']} (-{price_info['discount_pct']}%)")

    return {
        **trainer,
        "steam_appid": appid, "steam_name": steam_name,
        "steam_url": f"https://store.steampowered.com/app/{appid}/",
        "deck_compat": deck_compat,
        **price_info, **reviews,
        "last_fetched": now,
        "_price_updated_at": now if price_info.get("price_idr") is not None else None,
        "_country_code": config.country_code,
    }


def process_new_trainers(trainers: list[dict], config: Config) -> list[dict]:
    """For each NEW trainer, do full Steam lookup concurrently."""
    if not trainers:
        return []

    results = []
    with ThreadPoolExecutor(max_workers=config.max_workers) as executor:
        futures = {executor.submit(_process_single_new_trainer, t, config): t for t in trainers}
        with tqdm(total=len(trainers), desc="  New games", unit="game") as pbar:
            for future in as_completed(futures):
                result = future.result()
                results.append(result)
                pbar.update(1)

    return results


def _refresh_single_price(cached_entry: dict, config: Config) -> dict:
    """Refresh price for a single cached entry. Skips if price data is fresh (within TTL)
    AND the country code hasn't changed."""
    now = datetime.now()
    now_iso = now.isoformat()

    # Force refresh if country code changed since last fetch
    cached_country = cached_entry.get("_country_code")
    country_changed = cached_country != config.country_code

    # Skip price refresh if data is fresh AND country hasn't changed
    if not country_changed:
        price_updated = cached_entry.get("_price_updated_at")
        if price_updated:
            try:
                price_time = datetime.fromisoformat(price_updated)
                if now - price_time < timedelta(hours=config.cache_price_ttl_hours):
                    return cached_entry  # still fresh, skip API call
            except (ValueError, TypeError):
                pass  # invalid timestamp, refresh anyway

    appid = cached_entry.get("steam_appid")
    if not appid:
        return cached_entry

    app_details = get_steam_app_details(appid, config)
    if app_details:
        price_info = extract_price_info(app_details, config)
        cached_entry.update(price_info)
        cached_entry["_price_updated_at"] = now_iso
        cached_entry["_country_code"] = config.country_code

    cached_entry["last_fetched"] = now_iso
    return cached_entry


def refresh_prices(cached_results: list[dict], config: Config) -> list[dict]:
    """Refresh price/discount data for cached entries concurrently."""
    if not cached_results:
        return cached_results

    has_appid = [r for r in cached_results if r.get("steam_appid")]
    no_appid = [r for r in cached_results if not r.get("steam_appid")]

    refreshed = []
    with ThreadPoolExecutor(max_workers=config.max_workers) as executor:
        futures = {executor.submit(_refresh_single_price, r, config): r for r in has_appid}
        with tqdm(total=len(has_appid), desc="  Refreshing prices", unit="game") as pbar:
            for future in as_completed(futures):
                result = future.result()
                refreshed.append(result)
                pbar.update(1)

    return refreshed + no_appid


# ─── Excel Output ────────────────────────────────────────────

def write_excel(results: list[dict], output_path: Path, config: Config):
    """Write results to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FLiNG + Steam Deck"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B2838", end_color="1B2838", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    verified_fill = PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid")
    playable_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    unsupported_fill = PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid")
    sale_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    price_fmt = config.price_number_format

    headers = [
        "Game Name", "Trainer Date", "Options", "Last Updated",
        "Steam Name", "Steam App ID", "Deck Compatibility",
        config.price_header, "Original Price", "Discount %", "On Sale",
        "Rating", "Rating %", "Total Reviews",
        "Trainer URL", "Steam URL",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    deck_order = {"Verified": 0, "Playable": 1, "Unknown": 2, "Unsupported": 3, "Not Found": 4}
    sorted_results = sorted(results, key=lambda r: (
        deck_order.get(r.get("deck_compat", "Unknown"), 5),
        -(r.get("positive_pct", 0)),
    ))

    for row_idx, r in enumerate(sorted_results, 2):
        ws.cell(row=row_idx, column=1, value=r.get("game_name", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=r.get("trainer_date_str", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=r.get("options_count")).border = thin_border
        ws.cell(row=row_idx, column=4, value=r.get("last_updated", "")).border = thin_border
        ws.cell(row=row_idx, column=5, value=r.get("steam_name", "")).border = thin_border
        ws.cell(row=row_idx, column=6, value=r.get("steam_appid")).border = thin_border

        deck_cell = ws.cell(row=row_idx, column=7, value=r.get("deck_compat", "Unknown"))
        deck_cell.border = thin_border
        deck_cell.alignment = Alignment(horizontal="center")
        if r.get("deck_compat") == "Verified":
            deck_cell.fill = verified_fill
        elif r.get("deck_compat") == "Playable":
            deck_cell.fill = playable_fill
        elif r.get("deck_compat") == "Unsupported":
            deck_cell.fill = unsupported_fill

        price_val = r.get("price_idr")
        if price_val is not None and price_val > 0:
            price_cell = ws.cell(row=row_idx, column=8, value=price_val)
            price_cell.border = thin_border
            price_cell.number_format = price_fmt
        else:
            ws.cell(row=row_idx, column=8, value=r.get("price", "N/A")).border = thin_border

        orig_price = r.get("original_price_idr")
        if orig_price is not None and orig_price > 0:
            orig_cell = ws.cell(row=row_idx, column=9, value=orig_price)
            orig_cell.border = thin_border
            orig_cell.number_format = price_fmt
        else:
            ws.cell(row=row_idx, column=9, value="").border = thin_border

        disc = r.get("discount_pct", 0)
        disc_cell = ws.cell(row=row_idx, column=10, value=f"-{disc}%" if disc > 0 else "")
        disc_cell.border = thin_border
        if disc > 0:
            disc_cell.fill = sale_fill

        sale_cell = ws.cell(row=row_idx, column=11, value="YES" if r.get("on_sale") else "")
        sale_cell.border = thin_border
        if r.get("on_sale"):
            sale_cell.fill = sale_fill
            sale_cell.font = Font(bold=True, color="155724")

        ws.cell(row=row_idx, column=12, value=r.get("review_desc", "")).border = thin_border

        pct_cell = ws.cell(row=row_idx, column=13, value=r.get("positive_pct", 0))
        pct_cell.border = thin_border
        pct_cell.number_format = '0.0"%"'

        ws.cell(row=row_idx, column=14, value=r.get("total_reviews", 0)).border = thin_border
        ws.cell(row=row_idx, column=14).number_format = '#,##0'

        ws.cell(row=row_idx, column=15, value=r.get("trainer_url", "")).border = thin_border
        ws.cell(row=row_idx, column=16, value=r.get("steam_url", "")).border = thin_border

    col_widths = [35, 14, 8, 14, 35, 12, 18, 15, 15, 10, 8, 20, 10, 12, 50, 45]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sorted_results) + 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="FLiNG Trainer + Steam Deck Checker").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=3, column=1, value=f"Filter: Trainers from {config.min_year} onwards")
    ws2.cell(row=4, column=1, value=f"Region: {config.country_code} ({config.currency_code})")

    total = len(results)
    deck_verified = sum(1 for r in results if r.get("deck_compat") == "Verified")
    deck_playable = sum(1 for r in results if r.get("deck_compat") == "Playable")
    deck_unsupported = sum(1 for r in results if r.get("deck_compat") == "Unsupported")
    deck_unknown = sum(1 for r in results if r.get("deck_compat") in ("Unknown", "Not Found"))
    on_sale = sum(1 for r in results if r.get("on_sale"))
    free_games = sum(1 for r in results if r.get("price") == "Free")

    stats = [
        ("", ""),
        ("Total Trainers", total),
        ("Deck Verified", deck_verified),
        ("Deck Playable", deck_playable),
        ("Deck Unsupported", deck_unsupported),
        ("Unknown / Not Found", deck_unknown),
        ("", ""),
        ("Currently On Sale", on_sale),
        ("Free Games", free_games),
    ]
    for i, (label, val) in enumerate(stats, 6):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True) if label else Font()
        ws2.cell(row=i, column=2, value=val)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    print(f"\n✅ Excel saved to: {output_path}")


# ─── CLI ───────────────────────────────────────────────────────

def parse_args() -> Config:
    """Parse command-line arguments and return a Config object."""
    parser = argparse.ArgumentParser(
        description="FLiNG Trainer + Steam Deck Compatibility Checker",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Examples:
  %(prog)s                          Default: ID/IDR, trainers from 2024+
  %(prog)s --country US             Use USD pricing (United States)
  %(prog)s --country JP             Use JPY pricing (Japan)
  %(prog)s --year 2023             Include older trainers
  %(prog)s --no-cache              Fresh run, ignore existing cache
  %(prog)s --workers 2             Use 2 concurrent threads

Supported country codes for pricing:
  AR, AU, BR, CA, CN, EU, GB, ID, IN, JP, KR, MX, NO, NZ,
  PH, PL, RU, SA, SE, SG, TH, TR, TW, UA, US, ZA
""",
    )
    parser.add_argument("--year", type=int, default=2024,
                        help="Minimum trainer year (default: 2024)")
    parser.add_argument("--country", type=str, default="ID",
                        help="Steam store country code for pricing (default: ID)")
    parser.add_argument("--workers", type=int, default=4,
                        help="Number of concurrent threads (default: 4)")
    parser.add_argument("--delay", type=float, default=1.5,
                        help="Seconds between Steam API requests per game (default: 1.5)")
    parser.add_argument("--retries", type=int, default=3,
                        help="Max retry attempts for Steam API errors (default: 3)")
    parser.add_argument("--ttl-hours", type=int, default=24,
                        help="Skip price refresh if cached within N hours (default: 24)")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Output directory for Excel and cache (default: current dir)")
    parser.add_argument("--no-cache", action="store_true",
                        help="Ignore cache, fetch all data fresh")
    parser.add_argument("--verbose", action="store_true",
                        help="Show verbose/debug output")

    args = parser.parse_args()

    config = Config(
        min_year=args.year,
        country_code=args.country.upper(),
        request_delay=args.delay,
        max_workers=args.workers,
        max_retries=args.retries,
        cache_price_ttl_hours=args.ttl_hours,
        no_cache=args.no_cache,
        verbose=args.verbose,
        output_dir=Path(args.output_dir) if args.output_dir else None,
    )

    if config.country_code not in CURRENCY_MAP:
        print(f"⚠ Warning: Country code '{config.country_code}' not in CURRENCY_MAP.")
        print(f"  Prices will use Steam's raw format. Supported: {', '.join(sorted(CURRENCY_MAP.keys()))}")

    return config


# ─── Main ────────────────────────────────────────────────────

# Global session — created once, reused for all requests
SESSION = build_session()


def main():
    config = parse_args()

    print("=" * 60)
    print("FLiNG Trainer + Steam Deck Compatibility Checker")
    print("=" * 60)
    print(f"  Region: {config.country_code} ({config.currency_code})")
    print(f"  Workers: {config.max_workers} | Delay: {config.request_delay}s")

    # Load cache
    print(f"\n📦 Loading cache...")
    cache = load_cache(config)

    # Step 1: Scrape FLiNG (incremental — stops at cached entries)
    print(f"\n📋 Step 1: Scraping FLiNG Trainer (year >= {config.min_year})...")
    new_trainers, cached_results = scrape_fling_trainers(cache, config)

    if not new_trainers and not cached_results:
        print("No trainers found. Exiting.")
        return

    # Step 2a: Full Steam lookup for NEW trainers only
    new_results = []
    if new_trainers:
        print(f"\n🎮 Step 2a: Looking up {len(new_trainers)} NEW games on Steam...")
        new_results = process_new_trainers(new_trainers, config)
    else:
        print(f"\n🎮 Step 2a: No new games to look up!")

    # Step 2b: Refresh prices for cached entries
    print(f"\n💰 Step 2b: Refreshing prices for cached entries...")
    refreshed_cached = refresh_prices(cached_results, config)

    # Merge results
    all_results = new_results + refreshed_cached

    # Update cache with all results
    print(f"\n💾 Updating cache...")
    for r in all_results:
        url = r.get("trainer_url")
        if url:
            cache[url] = r
    save_cache(cache, config)

    # Step 3: Write Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = config.output_dir / f"fling_steam_deck_{timestamp}.xlsx"

    print(f"\n📊 Step 3: Writing Excel...")
    write_excel(all_results, output_path, config)

    # Summary
    deck_ok = [r for r in all_results if r.get("deck_compat") in ("Verified", "Playable")]
    on_sale = [r for r in deck_ok if r.get("on_sale")]

    print(f"\n{'=' * 60}")
    print(f"📈 Summary:")
    print(f"  Total trainers:        {len(all_results)}")
    print(f"  New (fetched):         {len(new_results)}")
    print(f"  Cached (refreshed):    {len(refreshed_cached)}")
    print(f"  Deck Compatible:       {len(deck_ok)} (Verified + Playable)")
    print(f"  On Sale (Deck OK):     {len(on_sale)}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()