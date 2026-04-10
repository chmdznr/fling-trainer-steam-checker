"""Excel output generation with formatting."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fling_checker.config import print, Config, DEFAULT_CURRENCY


def write_excel(results: list[dict], output_path: Path, config: Config):
    """Write results to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FLiNG + Steam Deck"

    # ─── Styles ──────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    playable_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    unsupported_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    sale_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # ─── Headers ─────────────────────────────────────────────
    currency = config.currency or DEFAULT_CURRENCY
    headers = [
        "Game Name", "Trainer Date", "Options", "Last Updated",
        "Steam Name", "Steam App ID",
        "Deck Compatibility",
        config.price_header,
        "Original Price", "Discount %", "On Sale",
        "Rating", "Rating %", "Total Reviews",
        "Trainer URL", "Steam URL",
    ]

    ws.append(headers)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ─── Sort: Verified > Playable > rest ────────────────────
    deck_order = {"Verified": 0, "Playable": 1, "Unsupported": 2, "Unknown": 3}
    results.sort(key=lambda r: (
        deck_order.get(r.get("deck_compat", "Unknown"), 3),
        -(r.get("positive_pct", 0) or 0),
    ))

    # ─── Data rows ──────────────────────────────────────────
    for row_idx, r in enumerate(results, 2):
        price_val = r.get("price_idr")
        orig_val = r.get("original_price_idr")

        row_data = [
            r.get("game_name", ""),
            r.get("trainer_date", ""),
            r.get("num_options", 0),
            r.get("version_info", ""),
            r.get("steam_name", ""),
            r.get("steam_appid", ""),
            r.get("deck_compat", "Unknown"),
            price_val if price_val is not None else r.get("price", "N/A"),
            orig_val if orig_val is not None else "",
            r.get("discount_pct", 0),
            "YES" if r.get("on_sale") else "NO",
            r.get("review_desc", ""),
            r.get("positive_pct", 0),
            r.get("total_reviews", 0),
            r.get("trainer_url", ""),
            r.get("steam_url", ""),
        ]
        ws.append(row_data)

        # Color-code deck compatibility
        deck_cell = ws.cell(row=row_idx, column=7)
        if deck_cell.value == "Verified":
            deck_cell.fill = verified_fill
        elif deck_cell.value == "Playable":
            deck_cell.fill = playable_fill
        elif deck_cell.value == "Unsupported":
            deck_cell.fill = unsupported_fill

        # Format price column
        price_cell = ws.cell(row=row_idx, column=8)
        if isinstance(price_val, (int, float)):
            price_cell.number_format = config.price_number_format

        # Color-code discount/on-sale
        discount_pct = r.get("discount_pct", 0)
        on_sale = r.get("on_sale", False)
        if on_sale or discount_pct > 0:
            for col in [10, 11]:  # Discount %, On Sale
                ws.cell(row=row_idx, column=col).fill = sale_fill

    # ─── Column widths ──────────────────────────────────────
    col_widths = {
        1: 35, 2: 12, 3: 8, 4: 18, 5: 35, 6: 12,
        7: 18, 8: 14, 9: 14, 10: 10, 11: 8,
        12: 22, 13: 10, 14: 12, 15: 45, 16: 45,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ─── Freeze top row + auto-filter ───────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ─── Summary sheet ───────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    deck_verified = sum(1 for r in results if r.get("deck_compat") == "Verified")
    deck_playable = sum(1 for r in results if r.get("deck_compat") == "Playable")
    deck_unsupported = sum(1 for r in results if r.get("deck_compat") == "Unsupported")
    deck_unknown = sum(1 for r in results if r.get("deck_compat") == "Unknown")
    on_sale = [r for r in results if r.get("on_sale") and r.get("deck_compat") in ("Verified", "Playable")]
    free_games = sum(1 for r in results if r.get("price") == "Free")

    summary_data = [
        ["Total Trainers", len(results)],
        ["Deck Verified", deck_verified],
        ["Deck Playable", deck_playable],
        ["Deck Unsupported", deck_unsupported],
        ["Deck Unknown", deck_unknown],
        ["Deck Compatible (Verified + Playable)", deck_verified + deck_playable],
        ["On Sale (Deck OK)", len(on_sale)],
        ["Free Games", free_games],
        ["Region", f"{config.country_code} ({config.currency_code})"],
        ["Generated", "N/A" if not results else results[0].get("last_fetched", "N/A")],

    ]

    for row in summary_data:
        ws2.append(row)

    # Style summary
    ws2.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws2.cell(row=1, column=2).font = Font(bold=True, size=12)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    print(f"\n✅ Excel saved to: {output_path}")